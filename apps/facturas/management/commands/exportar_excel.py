import openpyxl
from django.core.management.base import BaseCommand
from apps.facturas.models import Factura, LineaFactura


class Command(BaseCommand):
    help = "Exporta facturas a un fichero Excel con 3 hojas"

    def add_arguments(self, parser):
        parser.add_argument(
            'ruta',
            type=str,
            help='Ruta donde guardar el fichero .xlsx'
        )

    def handle(self, *args, **options):
        ruta = options['ruta']
        wb = openpyxl.Workbook()

        # Hoja 1: Facturas
        ws_facturas = wb.active
        ws_facturas.title = "Facturas"
        ws_facturas.append([
            'ID', 'Número', 'Fecha', 'Proveedor',
            'Categoría', 'Base imponible', 'IVA total', 'Total'
        ])
        for f in Factura.objects.select_related('proveedor', 'categoria').all():
            ws_facturas.append([
                f.pk,
                f.numero_factura,
                str(f.fecha_emision),
                str(f.proveedor),
                str(f.categoria) if f.categoria else '',
                float(f.base_imponible),
                float(f.iva_total),
                float(f.total),
            ])

        # Hoja 2: Líneas
        ws_lineas = wb.create_sheet("Líneas")
        ws_lineas.append([
            'ID', 'Factura ID', 'Concepto',
            'Cantidad', 'Precio unitario', 'IVA %', 'Total línea'
        ])
        for l in LineaFactura.objects.select_related('factura').all():
            ws_lineas.append([
                l.pk,
                l.factura.pk,
                l.concepto,
                float(l.cantidad),
                float(l.precio_unitario),
                l.iva_porcentaje,
                float(l.cantidad) * float(l.precio_unitario) * (1 + l.iva_porcentaje / 100),
            ])

        # Hoja 3: Resumen de IVA
        ws_iva = wb.create_sheet("Resumen IVA")
        ws_iva.append(['Tipo IVA', 'Base imponible', 'IVA soportado'])
        tipos_iva = [21, 10, 4, 0]
        for tipo in tipos_iva:
            lineas_tipo = LineaFactura.objects.filter(iva_porcentaje=tipo)
            base = sum(float(l.cantidad) * float(l.precio_unitario) for l in lineas_tipo)
            iva = sum(float(l.cantidad) * float(l.precio_unitario) * tipo / 100 for l in lineas_tipo)
            etiqueta = f"{tipo}%" if tipo > 0 else "Exento"
            ws_iva.append([etiqueta, base, iva])

        wb.save(ruta)
        self.stdout.write(f"Excel guardado en {ruta}")
